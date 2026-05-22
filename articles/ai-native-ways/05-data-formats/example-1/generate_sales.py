#!/usr/bin/env python3
"""100 個のサンプル売上 Excel (.xlsx) を生成する。

100 店舗 × 1 ヶ月 = 100 ファイル。各ファイル 100 行 = 合計 10,000 件の売上データ。
書式(セル幅、ヘッダ太字、色付き)も入れて、現実の業務 Excel に近づける。
これをあとで CSV に変換し、サイズ比とトークン比、集計時間を測る。
"""
from __future__ import annotations

import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42)

OUT = Path(__file__).parent / "xlsx"
OUT.mkdir(exist_ok=True)

ITEMS = [
    ("キャベツ", 180), ("玉葱", 95), ("にんじん", 120), ("じゃがいも", 80),
    ("トマト", 250), ("きゅうり", 90), ("レタス", 200), ("白菜", 150),
    ("大根", 130), ("ピーマン", 110), ("ナス", 140), ("ズッキーニ", 220),
]
STORES = [f"S{n:03d}" for n in range(1, 101)]


def make_workbook(store: str, month: int) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{store}-2026-{month:02d}"

    # ヘッダ(太字 + 色付き)
    headers = ["日付", "店舗", "商品", "数量", "単価", "売上"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5F8F")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # データ行(100 行)
    start = date(2026, month, 1)
    for row in range(2, 102):
        d = start + timedelta(days=random.randint(0, 27))
        item, base_price = random.choice(ITEMS)
        qty = random.randint(1, 50)
        price = base_price + random.randint(-20, 30)
        total = qty * price
        ws.cell(row=row, column=1, value=d.isoformat())
        ws.cell(row=row, column=2, value=store)
        ws.cell(row=row, column=3, value=item)
        ws.cell(row=row, column=4, value=qty)
        ws.cell(row=row, column=5, value=price)
        ws.cell(row=row, column=6, value=total)

    # 列幅
    for col, width in zip("ABCDEF", [12, 8, 14, 8, 8, 10]):
        ws.column_dimensions[col].width = width

    return wb


def main():
    month = 4  # 全ファイル 2026 年 4 月分(店舗ごと)
    for store in STORES:
        wb = make_workbook(store, month)
        path = OUT / f"{store}-2026-{month:02d}.xlsx"
        wb.save(path)
    files = list(OUT.glob("*.xlsx"))
    total = sum(p.stat().st_size for p in files)
    print(f"  生成完了: {len(files)} ファイル")
    print(f"  合計サイズ: {total:,} bytes ({total / 1024:.1f} KB)")


if __name__ == "__main__":
    main()
